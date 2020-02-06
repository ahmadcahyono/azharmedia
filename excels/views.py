from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.forms.models import model_to_dict
import django.http
from django.core import serializers
from django.db.models import F
from django.forms.models import model_to_dict
from django.core.files.base import ContentFile
import os
import openpyxl
from openpyxl import Workbook, load_workbook


from excels.models import (
    JenisData,

)

from excels.forms import (FrmJenisData)

from master.models import (
    Pelanggan, Suplier,

)
import uuid
import base64

from base64 import decodestring


from config.models import (
     Kabupaten as kab,
)

from django.views import View
from base64 import b64encode, b64decode


from django.views.generic.base import (
    RedirectView, 
    TemplateView,
    # View,
)

@login_required
def index(request):
  return render(request, 'excel/pages/home.html')


@login_required   
def latihan(request):
    if "POST" == request.method and request.FILES['browse']:
        myfile = request.FILES['browse']
        info = "orang pinggiran"
        image = i.open(myfile)
        w, h = image.size
        file_path = os.path.dirname(os.path.abspath(myfile.name))
        # image = image.rezise((int(w)/2,int(h)/2), i.ANTIALIAS)
        #image_path = settings.STATIC_URL + "/users/" + myfile.name + "." 
        # + myfile.name + "." + myfile.ext 
        # "uploads/" + myfile.name
        image_path = ""
        if os.path.exists(image_path):
            info = "file exist"
            uploaded_file_url = image_path
        else:
            fs = FileSystemStorage()
            info = "ok"
            filename = fs.save("uploads/" + myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
        
        return render(request, 'excel/pages/latihan.html', {'info'  : info, 'file_url' : uploaded_file_url})
        
    else :
        image_path = settings.STATIC_URL + "images/users/avatar-1.jpg"
        uploaded_file_url = image_path
        return render(request, 'excel/pages/latihan.html', {'info' : 'awal','file_url' :  uploaded_file_url} )



def load_excel(request):
    
    if "GET" == request.method:
         return render(request, 'excel/pages/load_excels.html', {})
    else :    
        excel_file = request.FILES["excel_file"]
        excel_path = os.path.dirname(os.path.abspath(excel_file.name))
        count_rows = 0
        count_cols = 0
        cols = 0
   
        
        wb = openpyxl.load_workbook(excel_file)
        # getting a particular sheet by name out of many sheets
        sheets = wb.sheetnames
        jumlah = len(sheets)
        excel_info = ({'sheets' : jumlah})
        sheetsku = []
        
        for n in range(jumlah) :
          sheetsku.append(sheets[n])
          
          

        worksheet = wb.active
        
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            count_rows= count_rows + 1
            for cell in row:
                row_data.append(str(cell.value))
                count_cols = count_cols + 1 
            excel_data.append(row_data)
        cols = count_cols
        excel_info['rows'] = count_rows 
        excel_info['cols'] : count_cols/count_rows  
        return render(request, 'excel/pages/load_excels.html', 
            {"excel_data":excel_data,
            "sheets":sheetsku, 
            'jml':jumlah,
            'cols' : count_cols/count_rows, 
            'excel_info': excel_info,
            'excel_path' : excel_path, 
                })
    # else:
    #     return render(request, 'pages/excel.html', {})  


@login_required
def edit_jenisdata(request, update_id):
  jenisdata_update = JenisData.objects.get(id=update_id)
  data = {
    'nama_jd' : jenisdata_update.nama_jd,
    'tabel_jd' : jenisdata_update.tabel_jd,
  }
  form = FrmJenisData(request.POST or None, initial=data, instance=jenisdata_update)
  if request.method == 'POST':
    if form.is_valid():
      form.save()
      return redirect('proses:jenisdata')

  list_jd = JenisData.objects.all()
  context = {
    'title' : 'Jenis Data',
    'list_d' : list_jd,
    'form_d' : form,
    'tombol' : 'Edit',
    'pesan'  : 'Perubahan data',
  }       
  return render(request, 'excel/pages/jenisdata.html',context)


@login_required
def delete_jenisdata(request, delete_id):
  jd_edit = JenisData.objects.filter(id=delete_id).delete()
  return redirect('proses:jenisdata')
  


@login_required
def jenisdata(request):
  list_jd = JenisData.objects.all()
  form_jd = FrmJenisData(request.POST or None)
  if request.method == 'POST':
    if form_jd.is_valid():
      form_jd.save()
      list_jd = JenisData.objects.all()
      context = {
        'title' : 'Jenis Data',
        'list_d' : list_jd,
        'form_d' : form_jd,
        'tombol' : 'Simpan',
      }       
      return render(request, 'excel/pages/jenisdata.html',context)
    
  context = {
    'title' : 'Jenis Data',
    'list_d' : list_jd,
    'form_d' : form_jd,
    'tombol' : 'Simpan',
  }       
    
  return render(request, 'excel/pages/jenisdata.html', context)


@login_required
def criteria(request):
   hasil = dict()
   hasil['jenisdata'] = JenisData.objects.all().values('nama_jd')
   hasil['pelanggan'] = Pelanggan.objects.all().values('id', 'nama_pelanggan').annotate(nama=F('nama_pelanggan'))
   hasil['suplier'] = Suplier.objects.all().values('id', 'nama_suplier').annotate(nama=F('nama_suplier'))
   return render(request, 'excel/pages/kriteria.html', hasil)

@login_required
def dataprocess(request):
   return render(request, 'excel/pages/dataproses.html')



@login_required
def kalimat(request):
   return render(request, 'excel/pages/kalimat.html')



def databaseku(mydata):
    dataku = {}
    if(mydata.lower()=="agen"):
      mode='Agen'
      dataku1 = Pelanggan.objects.all().annotate(nama=F('nama_pelanggan'))
      dataku = Pelanggan.objects.all().values('id', 'nama_pelanggan').annotate(nama=F('nama_pelanggan'))
    elif(mydata.lower()=="suplier"):
      mode='Suplier'
      dataku = Suplier.objects.all().annotate(nama=F('nama_suplier'))
    elif(mydata.lower()=="bank"):
      mode='Bank'
      dataku = Kamus.objects.all().annotate(nama=F('frasa'))
    elif(mydata.lower()=="karyawan"):
      mode='Karyawan'
      dataku = Karyawan.objects.all().annotate(nama=F('nama_karyawan'))
    elif(mydata.lower()=="produk"):
      mode='Produk'
      dataku = Produk.objects.all().annotate(nama=F('nama_produk'))
    return dataku



@csrf_exempt
def gantitabel(request):
    dataku = {}
    dataku1 = dict()
    mode = ''
    if request.method == "POST":
      
      mydata = request.POST.get('asing')
      dataku = databaseku(mydata)
      # ser_obj = serializers.serialize('json', dataku)
      # struct = json.loads(ser_obj)
      # data = json.dumps(struct)
      
      baris = list()
      

      for isi in dataku:
        baris.append((isi['id'],  isi['nama']))
      hasilnya = json.dumps(baris)
      print(hasilnya)
      return HttpResponse(hasilnya, content_type='application/json')


"""
Came up with this for satchmo's downloadable product migration, 0001_split.
"""
def db_table_exists(table, cursor=None):
    from django.db import connection
    return table in connection.introspection.table_names()


@csrf_exempt
def getdata(request,table):
  dataku = {}
  baris = list()
  store = table.title()
  dataku = databaseku(table)
  for isi in dataku:
        baris.append((isi['id'],  isi['nama']))
      
  tamp = "excels_" + table
  hasil = db_table_exists(tamp)
  data = json.dumps(baris)
      
  # sadino = {'nama' : hasil, 'v' : vasa}
  return HttpResponse(data , content_type='applcation/json')


class CSRFExemptMixin(object):
    @csrf_exempt
    def dispatch(self, request, *args, **kwargs):
        return super(CSRFExemptMixin, self).dispatch(request, *args, **kwargs)



# @csrf_exempt
class AmbilData(TemplateView, CSRFExemptMixin):
  template_name = "excel/pages/kriteria/lat.html"
  pattern = "asoka"
  permanent = False




@csrf_exempt
def getjenisdata(request):
  dataku = {}
  baris = list()
  dataku = JenisData.objects.all().values('id', 'nama_jd');
  
  for isi in dataku:
        baris.append((isi['id'],  isi['nama_jd']))
  # sadino = {'nama' : hasil, 'v' : vasa}
  data = json.dumps(baris)
  return HttpResponse(data , content_type='application/json')



